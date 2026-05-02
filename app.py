import sqlite3
from flask import Flask, render_template, g, request, redirect, url_for, flash, session
from datetime import datetime, timedelta
from functools import wraps
import hashlib
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'super-secret-key-for-college-project'
app.config['DATABASE'] = 'database.db'


# --- ДЕКОРАТОР ДЛЯ ЗАЩИТЫ МАРШРУТОВ ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Пожалуйста, войдите в систему', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


# --- Работа с БД ---
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


def init_db():
    """СОЗДАЕТ БД С НУЛЯ"""
    with app.app_context():
        db = get_db()

        # Удаляем старые таблицы если есть
        db.execute('DROP TABLE IF EXISTS history')
        db.execute('DROP TABLE IF EXISTS equipment')
        db.execute('DROP TABLE IF EXISTS users')

        # Создаем таблицу equipment
        db.execute('''
            CREATE TABLE equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                type TEXT NOT NULL,
                inventory_number TEXT UNIQUE NOT NULL,
                rfid_tag TEXT UNIQUE,
                location TEXT,
                status TEXT DEFAULT 'В работе',
                purchase_date TEXT,
                warranty_months INTEGER,
                next_check_date TEXT,
                responsible_person TEXT,
                last_maintenance TEXT,
                notes TEXT,
                price REAL
            )
        ''')

        # Создаем таблицу истории
        db.execute('''
            CREATE TABLE history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_id INTEGER NOT NULL,
                event_date TEXT DEFAULT CURRENT_TIMESTAMP,
                event_type TEXT,
                description TEXT,
                user TEXT,
                FOREIGN KEY (equipment_id) REFERENCES equipment (id) ON DELETE CASCADE
            )
        ''')

        # Создаем таблицу пользователей
        db.execute('''
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                email TEXT,
                phone TEXT,
                full_name TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        db.commit()

        # Добавляем тестовые данные
        add_sample_data()
        print("База данных успешно создана!")


def add_sample_data():
    """Тестовые данные"""
    db = get_db()

    # Тестовое оборудование
    samples = [
        ('Ноутбук HP ProBook', 'Ноутбук', 'INV-001', 'RFID-001', 'Кабинет 301', 'В работе',
         '2024-01-15', 24, '2026-01-15', 'Иванов И.И.', '2024-10-01', 'Отличное состояние', 85000),
        ('Принтер Canon', 'Принтер', 'INV-002', 'RFID-002', 'Кабинет 302', 'На ремонте',
         '2023-11-20', 12, '2024-11-20', 'Петров П.П.', '2024-09-15', 'Требуется ремонт', 25000),
        ('Проектор Epson', 'Проектор', 'INV-003', 'RFID-003', 'Актовый зал', 'В работе',
         '2024-03-10', 24, '2026-03-10', 'Сидоров С.С.', '2024-08-20', 'Лампа новая', 45000),
        ('МФУ Xerox', 'МФУ', 'INV-004', 'RFID-004', 'Кабинет 305', 'В работе',
         '2023-08-05', 18, '2025-02-05', 'Иванов И.И.', '2024-09-01', 'Сканер работает', 35000),
        ('Сервер Dell', 'Сервер', 'INV-005', 'RFID-005', 'Серверная', 'В работе',
         '2024-02-20', 36, '2027-02-20', 'Петров П.П.', '2024-10-10', 'Резервирование', 250000),
        ('Монитор Samsung', 'Монитор', 'INV-006', 'RFID-006', 'Кабинет 301', 'В работе',
         '2024-01-10', 24, '2026-01-10', 'Иванов И.И.', '2024-09-20', 'Без дефектов', 18000),
        ('Клавиатура Logitech', 'Клавиатура', 'INV-007', 'RFID-007', 'Кабинет 302', 'Списано',
         '2022-05-15', 12, '2023-05-15', 'Петров П.П.', '2023-04-01', 'Износ', 3500),
    ]

    for item in samples:
        try:
            db.execute('''
                INSERT INTO equipment 
                (name, type, inventory_number, rfid_tag, location, status, 
                 purchase_date, warranty_months, next_check_date, responsible_person,
                 last_maintenance, notes, price)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', item)
        except:
            pass

    # Добавляем тестового пользователя (логин: admin, пароль: admin123)
    db.execute('''
        INSERT OR IGNORE INTO users (username, password, email, phone, full_name)
        VALUES (?, ?, ?, ?, ?)
    ''', ('admin', hashlib.sha256('admin123'.encode()).hexdigest(), 'admin@surpk.ru', '+7(999)123-45-67',
          'Иванов Иван Иванович'))

    db.commit()


@app.context_processor
def inject_now():
    return {'now': datetime.now()}


@app.context_processor
def inject_user():
    return {'session': session}


# --- АВТОРИЗАЦИЯ ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = hashlib.sha256(request.form['password'].encode()).hexdigest()

        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username=? AND password=?', (username, password)).fetchone()

        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user['full_name']
            session['email'] = user['email']
            session['phone'] = user['phone']
            flash('Вход выполнен успешно!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверный логин или пароль', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Вы вышли из системы', 'success')
    return redirect(url_for('login'))


@app.route('/profile')
@login_required
def profile():
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    return render_template('profile.html', user=user)


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old = hashlib.sha256(request.form['old_password'].encode()).hexdigest()
        new = hashlib.sha256(request.form['new_password'].encode()).hexdigest()

        db = get_db()
        user = db.execute('SELECT * FROM users WHERE id=? AND password=?', (session['user_id'], old)).fetchone()

        if user:
            db.execute('UPDATE users SET password=? WHERE id=?', (new, session['user_id']))
            db.commit()
            flash('Пароль успешно изменён!', 'success')
            return redirect(url_for('profile'))
        else:
            flash('Неверный текущий пароль', 'error')

    return render_template('change_password.html')


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email_or_phone = request.form['email_or_phone']
        # Здесь можно добавить логику отправки кода на email/телефон
        flash('Инструкция по восстановлению отправлена (демо-режим)', 'success')
    return render_template('forgot_password.html')


# --- ГЛАВНАЯ ---
@app.route('/')
@login_required
def index():
    db = get_db()

    total = db.execute('SELECT COUNT(*) as c FROM equipment').fetchone()['c']
    in_repair = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="На ремонте"').fetchone()['c']

    soon_check = db.execute('''
        SELECT COUNT(*) as c FROM equipment 
        WHERE next_check_date IS NOT NULL 
        AND julianday(next_check_date) - julianday('now') < 30
        AND julianday(next_check_date) - julianday('now') > 0
    ''').fetchone()['c']

    overdue = db.execute('''
        SELECT COUNT(*) as c FROM equipment 
        WHERE next_check_date IS NOT NULL 
        AND julianday('now') - julianday(next_check_date) > 0
    ''').fetchone()['c']

    recent = db.execute('SELECT * FROM equipment ORDER BY id DESC LIMIT 5').fetchall()
    types_stats = db.execute('SELECT type, COUNT(*) as c FROM equipment GROUP BY type ORDER BY c DESC').fetchall()

    return render_template('index.html', total=total, in_repair=in_repair,
                           soon_check=soon_check, overdue=overdue, recent=recent, types_stats=types_stats)


# --- СПИСОК ОБОРУДОВАНИЯ ---
@app.route('/equipment')
@login_required
def equipment_list():
    db = get_db()

    status = request.args.get('status', '')
    type_filter = request.args.get('type', '')
    search = request.args.get('search', '')

    query = 'SELECT * FROM equipment WHERE 1=1'
    params = []

    if status:
        query += ' AND status = ?'
        params.append(status)
    if type_filter:
        query += ' AND type = ?'
        params.append(type_filter)
    if search:
        query += ' AND (name LIKE ? OR inventory_number LIKE ? OR responsible_person LIKE ?)'
        params.append(f'%{search}%')
        params.append(f'%{search}%')
        params.append(f'%{search}%')

    query += ' ORDER BY name'

    equipments = db.execute(query, params).fetchall()
    types = db.execute('SELECT DISTINCT type FROM equipment ORDER BY type').fetchall()

    return render_template('equipment.html', equipments=equipments, types=types,
                           current_status=status, current_type=type_filter, search_query=search)


# --- ДЕТАЛИ ---
@app.route('/equipment/<int:equip_id>')
@login_required
def equipment_detail(equip_id):
    db = get_db()
    equipment = db.execute('SELECT * FROM equipment WHERE id = ?', (equip_id,)).fetchone()

    if not equipment:
        flash('Оборудование не найдено', 'error')
        return redirect(url_for('equipment_list'))

    history = db.execute('SELECT * FROM history WHERE equipment_id = ? ORDER BY event_date DESC LIMIT 10',
                         (equip_id,)).fetchall()

    return render_template('equipment_detail.html', equipment=equipment, history=history)


# --- ДОБАВЛЕНИЕ ---
@app.route('/equipment/add', methods=['GET', 'POST'])
@login_required
def add_equipment():
    if request.method == 'POST':
        try:
            name = request.form['name'].strip()
            eq_type = request.form['type'].strip()
            inv_num = request.form['inventory_number'].strip()
            rfid = request.form.get('rfid_tag', '').strip() or None
            location = request.form.get('location', '').strip() or None
            status = request.form.get('status', 'В работе')
            responsible = request.form.get('responsible_person', '').strip() or None
            purchase_date = request.form.get('purchase_date', '').strip() or None
            warranty = request.form.get('warranty_months', '').strip() or None
            price = request.form.get('price', '').strip() or None
            notes = request.form.get('notes', '').strip() or None

            if not name or not inv_num or not eq_type:
                flash('Заполните обязательные поля', 'error')
                return redirect(url_for('add_equipment'))

            next_check = None
            if purchase_date and warranty:
                try:
                    purchase = datetime.strptime(purchase_date, '%Y-%m-%d')
                    next_check = (purchase + timedelta(days=30 * int(warranty))).strftime('%Y-%m-%d')
                except:
                    next_check = None

            db = get_db()

            existing = db.execute('SELECT id FROM equipment WHERE inventory_number = ?', (inv_num,)).fetchone()
            if existing:
                flash('Инвентарный номер уже существует', 'error')
                return redirect(url_for('add_equipment'))

            cursor = db.execute('''
                INSERT INTO equipment 
                (name, type, inventory_number, rfid_tag, location, status, 
                 responsible_person, purchase_date, warranty_months, next_check_date, notes, price)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, eq_type, inv_num, rfid, location, status, responsible,
                  purchase_date, warranty, next_check, notes, price))

            db.execute('INSERT INTO history (equipment_id, event_type, description, user) VALUES (?, ?, ?, ?)',
                       (cursor.lastrowid, 'Добавление', f'Добавлено: {name}', session.get('username', 'admin')))
            db.commit()

            flash('Оборудование успешно добавлено!', 'success')
            return redirect(url_for('equipment_list'))

        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'error')
            return redirect(url_for('add_equipment'))

    return render_template('add_equipment.html')


# --- РЕДАКТИРОВАНИЕ ---
@app.route('/equipment/<int:equip_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_equipment(equip_id):
    db = get_db()

    if request.method == 'POST':
        try:
            name = request.form['name'].strip()
            eq_type = request.form['type'].strip()
            location = request.form.get('location', '').strip() or None
            status = request.form.get('status', 'В работе')
            responsible = request.form.get('responsible_person', '').strip() or None
            notes = request.form.get('notes', '').strip() or None

            db.execute('''
                UPDATE equipment 
                SET name=?, type=?, location=?, status=?, responsible_person=?, notes=?
                WHERE id=?
            ''', (name, eq_type, location, status, responsible, notes, equip_id))

            db.execute('INSERT INTO history (equipment_id, event_type, description, user) VALUES (?, ?, ?, ?)',
                       (equip_id, 'Редактирование', 'Данные обновлены', session.get('username', 'admin')))
            db.commit()

            flash('Изменения сохранены', 'success')
            return redirect(url_for('equipment_detail', equip_id=equip_id))

        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'error')

    equipment = db.execute('SELECT * FROM equipment WHERE id = ?', (equip_id,)).fetchone()
    return render_template('edit_equipment.html', equipment=equipment)


# --- УДАЛЕНИЕ ---
@app.route('/equipment/<int:equip_id>/delete', methods=['POST'])
@login_required
def delete_equipment(equip_id):
    db = get_db()
    equipment = db.execute('SELECT name FROM equipment WHERE id = ?', (equip_id,)).fetchone()

    if equipment:
        db.execute('DELETE FROM equipment WHERE id = ?', (equip_id,))
        db.commit()
        flash(f'Оборудование "{equipment["name"]}" удалено', 'success')

    return redirect(url_for('equipment_list'))


# --- МЕСТОПОЛОЖЕНИЯ ---
@app.route('/locations')
@login_required
def locations():
    db = get_db()

    locations = db.execute('''
        SELECT location, COUNT(*) as count,
               GROUP_CONCAT(DISTINCT type) as types,
               GROUP_CONCAT(DISTINCT status) as statuses
        FROM equipment 
        WHERE location IS NOT NULL AND location != ''
        GROUP BY location
        ORDER BY location
    ''').fetchall()

    buildings = {}
    for loc in locations:
        building = loc['location'].split()[0] if loc['location'] else 'Другое'
        buildings[building] = buildings.get(building, 0) + loc['count']

    total_locations = len(locations)
    total_equipment = db.execute('SELECT COUNT(*) as c FROM equipment').fetchone()['c']

    return render_template('locations.html', locations=locations, buildings=buildings,
                           total_locations=total_locations, total_equipment=total_equipment)


# --- СОТРУДНИКИ ---
@app.route('/employees')
@login_required
def employees():
    db = get_db()

    employees = db.execute('''
        SELECT responsible_person, 
               COUNT(*) as total,
               SUM(CASE WHEN status = 'В работе' THEN 1 ELSE 0 END) as working,
               SUM(CASE WHEN status = 'На ремонте' THEN 1 ELSE 0 END) as repair,
               SUM(CASE WHEN status = 'На складе' THEN 1 ELSE 0 END) as storage,
               SUM(CASE WHEN status = 'Списано' THEN 1 ELSE 0 END) as written_off,
               GROUP_CONCAT(DISTINCT type) as types,
               GROUP_CONCAT(DISTINCT location) as locations
        FROM equipment 
        WHERE responsible_person IS NOT NULL AND responsible_person != ''
        GROUP BY responsible_person
        ORDER BY responsible_person
    ''').fetchall()

    total_employees = len(employees)
    total_equipment = db.execute('SELECT COUNT(*) as c FROM equipment').fetchone()['c']

    return render_template('employees.html', employees=employees,
                           total_employees=total_employees, total_equipment=total_equipment)


# --- ОТЧЕТЫ ---
@app.route('/reports')
@login_required
def reports():
    db = get_db()

    status_stats = db.execute('''
        SELECT status, COUNT(*) as count 
        FROM equipment 
        GROUP BY status
    ''').fetchall()

    type_stats = db.execute('''
        SELECT type, COUNT(*) as count 
        FROM equipment 
        GROUP BY type 
        ORDER BY count DESC
    ''').fetchall()

    monthly_stats = db.execute('''
        SELECT strftime('%Y-%m', purchase_date) as month,
               COUNT(*) as count
        FROM equipment 
        WHERE purchase_date IS NOT NULL
        GROUP BY month
        ORDER BY month DESC
        LIMIT 6
    ''').fetchall()

    no_responsible = db.execute('''
        SELECT COUNT(*) as c 
        FROM equipment 
        WHERE responsible_person IS NULL OR responsible_person = ''
    ''').fetchone()['c']

    no_location = db.execute('''
        SELECT COUNT(*) as c 
        FROM equipment 
        WHERE location IS NULL OR location = ''
    ''').fetchone()['c']

    expiring_soon = db.execute('''
        SELECT COUNT(*) as c 
        FROM equipment 
        WHERE next_check_date IS NOT NULL 
        AND julianday(next_check_date) - julianday('now') < 30
        AND julianday(next_check_date) - julianday('now') > 0
    ''').fetchone()['c']

    total_price = db.execute('SELECT SUM(price) as total FROM equipment').fetchone()['total'] or 0
    total = db.execute('SELECT COUNT(*) as c FROM equipment').fetchone()['c']

    return render_template('reports.html',
                           status_stats=status_stats,
                           type_stats=type_stats,
                           monthly_stats=monthly_stats,
                           no_responsible=no_responsible,
                           no_location=no_location,
                           expiring_soon=expiring_soon,
                           total_price=total_price,
                           total=total)


# --- ЭКСПОРТ В EXCEL ---
@app.route('/export/excel')
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        db = get_db()
        equipments = db.execute('SELECT * FROM equipment ORDER BY name').fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Оборудование_СПК"

        headers = ['Инвентарный номер', 'Наименование', 'Тип', 'Статус',
                   'Местоположение', 'Ответственный', 'RFID метка',
                   'Дата покупки', 'Гарантия', 'След. проверка', 'Стоимость']

        header_font = Font(bold=True, color="FFFFFF", size=11, name='Arial')
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row, item in enumerate(equipments, 2):
            ws.cell(row=row, column=1, value=str(item['inventory_number']) if item['inventory_number'] else '')
            ws.cell(row=row, column=2, value=str(item['name']) if item['name'] else '')
            ws.cell(row=row, column=3, value=str(item['type']) if item['type'] else '')
            ws.cell(row=row, column=4, value=str(item['status']) if item['status'] else '')
            ws.cell(row=row, column=5, value=str(item['location']) if item['location'] else '')
            ws.cell(row=row, column=6, value=str(item['responsible_person']) if item['responsible_person'] else '')
            ws.cell(row=row, column=7, value=str(item['rfid_tag']) if item['rfid_tag'] else '')
            ws.cell(row=row, column=8, value=str(item['purchase_date']) if item['purchase_date'] else '')
            ws.cell(row=row, column=9, value=str(item['warranty_months']) if item['warranty_months'] else '')
            ws.cell(row=row, column=10, value=str(item['next_check_date']) if item['next_check_date'] else '')
            ws.cell(row=row, column=11, value=float(item['price']) if item['price'] else 0)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        from flask import send_file
        return send_file(
            excel_file,
            download_name=f'oborudovanie_spk_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Ошибка при экспорте в Excel: {str(e)}', 'error')
        return redirect(url_for('reports'))


# --- ЭКСПОРТ В PDF ---
@app.route('/export/pdf')
@login_required
def export_pdf():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import os

        font_name = 'Helvetica'
        try:
            font_path = "C:/Windows/Fonts/arial.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Arial', font_path))
                font_name = 'Arial'
        except:
            pass

        db = get_db()
        equipments = db.execute('SELECT * FROM equipment ORDER BY name').fetchall()

        total = len(equipments)
        in_repair = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="На ремонте"').fetchone()['c']
        working = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="В работе"').fetchone()['c']

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                leftMargin=15 * mm, rightMargin=15 * mm,
                                topMargin=20 * mm, bottomMargin=15 * mm)
        elements = []

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.HexColor('#1E40AF'),
            alignment=1,
            spaceAfter=20,
            fontName=font_name
        )

        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontSize=9,
            fontName=font_name
        )

        title_text = "Отчет по оборудованию Сургутского политехнического колледжа"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 5))

        date_text = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        elements.append(Paragraph(date_text, normal_style))
        elements.append(Spacer(1, 15))

        stats_data = [
            ['Показатель', 'Количество'],
            ['Всего оборудования', str(total)],
            ['В работе', str(working)],
            ['На ремонте', str(in_repair)]
        ]

        stats_table = Table(stats_data, colWidths=[80 * mm, 40 * mm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))

        data = [['Инв. №', 'Наименование', 'Тип', 'Статус', 'Местоположение', 'Ответственный']]

        for item in equipments:
            inv_num = str(item['inventory_number']) if item['inventory_number'] else '-'
            name = str(item['name'])[:30] if item['name'] else '-'
            if len(str(item['name'])) > 30:
                name = name + '...'
            eq_type = str(item['type']) if item['type'] else '-'
            status = str(item['status']) if item['status'] else '-'
            location = str(item['location'])[:15] if item['location'] else '-'
            responsible = str(item['responsible_person'])[:15] if item['responsible_person'] else '-'

            data.append([inv_num, name, eq_type, status, location, responsible])

        col_widths = [40 * mm, 55 * mm, 30 * mm, 35 * mm, 35 * mm, 40 * mm]
        equipment_table = Table(data, colWidths=col_widths, repeatRows=1)

        equipment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))

        elements.append(equipment_table)

        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.grey,
            fontName=font_name
        )
        footer_text = "Отчет сгенерирован автоматически в системе АИС 'Цифровой паспорт оборудования'"
        elements.append(Paragraph(footer_text, footer_style))

        doc.build(elements)
        buffer.seek(0)

        from flask import send_file
        return send_file(
            buffer,
            download_name=f'otchet_spk_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            as_attachment=True,
            mimetype='application/pdf'
        )
    except Exception as e:
        flash(f'Ошибка при экспорте в PDF: {str(e)}', 'error')
        return redirect(url_for('reports'))


# --- ПЕЧАТЬ ОТЧЕТА ---
@app.route('/print-report')
@login_required
def print_report():
    db = get_db()

    equipments = db.execute('SELECT * FROM equipment ORDER BY name').fetchall()
    total = len(equipments)
    in_repair = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="На ремонте"').fetchone()['c']
    working = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="В работе"').fetchone()['c']
    storage = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="На складе"').fetchone()['c']
    written_off = db.execute('SELECT COUNT(*) as c FROM equipment WHERE status="Списано"').fetchone()['c']

    return render_template('print_report.html',
                           equipments=equipments,
                           total=total,
                           in_repair=in_repair,
                           working=working,
                           storage=storage,
                           written_off=written_off,
                           now=datetime.now())


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)